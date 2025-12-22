"""
Excel数据操作层

每个sheet存储一周的数据，sheet名格式: YYYY-WW (如 2024-51)
另有 recommender_stats sheet 存储推荐人统计
"""
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import config


class ExcelDB:
    """Excel数据库操作类"""
    
    # 股票数据列
    STOCK_COLUMNS = [
        'stock_name', 'market', 'stock_code', 'recommenders',
        'open_price', 'close_price', 'change_pct', 'status'
    ]
    
    # 推荐人统计列
    STATS_COLUMNS = [
        'name', 'total_count', 'win_count', 'win_rate', 'positive_weeks',
        'weeks_count', 'weekly_win_rate', 'total_return', 'avg_return',
        'score', 'rating', 'weekly_returns_json', 'details_json'
    ]
    
    def __init__(self, excel_path: str = None):
        self.excel_path = excel_path or config.EXCEL_FILE
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """确保Excel文件存在"""
        if not os.path.exists(self.excel_path):
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            wb = Workbook()
            # 创建推荐人统计sheet
            ws = wb.active
            ws.title = 'recommender_stats'
            for i, col in enumerate(self.STATS_COLUMNS, 1):
                ws.cell(row=1, column=i, value=col)
            # 创建raw_text sheet存储原始文本
            ws_raw = wb.create_sheet('raw_text')
            ws_raw.cell(row=1, column=1, value='year')
            ws_raw.cell(row=1, column=2, value='week')
            ws_raw.cell(row=1, column=3, value='raw_text')
            ws_raw.cell(row=1, column=4, value='recommender_messages')
            wb.save(self.excel_path)
    
    def _get_week_sheet_name(self, year: int, week: int) -> str:
        """获取周sheet名"""
        return f"{year}-{week:02d}"
    
    def _get_workbook(self) -> Workbook:
        """获取工作簿"""
        return load_workbook(self.excel_path)
    
    # ==================== 周数据操作 ====================
    
    def get_week_data(self, year: int, week: int) -> Dict[str, Any]:
        """获取指定周的数据"""
        wb = self._get_workbook()
        sheet_name = self._get_week_sheet_name(year, week)
        
        if sheet_name not in wb.sheetnames:
            return {'stocks': [], 'raw_text': '', 'recommender_messages': {}}
        
        ws = wb[sheet_name]
        stocks = []
        
        for row in range(2, ws.max_row + 1):
            stock = {}
            for i, col in enumerate(self.STOCK_COLUMNS, 1):
                value = ws.cell(row=row, column=i).value
                if col == 'recommenders' and value:
                    stock[col] = value.split(',')
                else:
                    stock[col] = value
            if stock.get('stock_name'):
                stocks.append(stock)
        
        # 获取原始文本
        raw_text, recommender_messages = self._get_raw_text(year, week)
        
        return {
            'stocks': stocks,
            'raw_text': raw_text,
            'recommender_messages': recommender_messages
        }
    
    def save_week_data(self, year: int, week: int, stocks: List[Dict], 
                       raw_text: str = '', recommender_messages: Dict = None):
        """保存周数据"""
        wb = self._get_workbook()
        sheet_name = self._get_week_sheet_name(year, week)
        
        # 删除已有sheet
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        # 创建新sheet
        ws = wb.create_sheet(sheet_name)
        
        # 写入表头
        for i, col in enumerate(self.STOCK_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col)
        
        # 写入数据
        for row_idx, stock in enumerate(stocks, 2):
            for col_idx, col in enumerate(self.STOCK_COLUMNS, 1):
                value = stock.get(col, '')
                if col == 'recommenders' and isinstance(value, list):
                    value = ','.join(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(self.excel_path)
        
        # 保存原始文本
        if raw_text:
            self._save_raw_text(year, week, raw_text, recommender_messages or {})
    
    def update_stock(self, year: int, week: int, stock_name: str, updates: Dict):
        """更新单只股票数据"""
        wb = self._get_workbook()
        sheet_name = self._get_week_sheet_name(year, week)
        
        if sheet_name not in wb.sheetnames:
            return False
        
        ws = wb[sheet_name]
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == stock_name:
                for col_idx, col in enumerate(self.STOCK_COLUMNS, 1):
                    if col in updates:
                        ws.cell(row=row, column=col_idx, value=updates[col])
                wb.save(self.excel_path)
                return True
        return False
    
    def get_all_weeks(self) -> List[Dict[str, int]]:
        """获取所有周列表"""
        wb = self._get_workbook()
        weeks = []
        for sheet_name in wb.sheetnames:
            if '-' in sheet_name and sheet_name not in ['recommender_stats', 'raw_text']:
                try:
                    year, week = sheet_name.split('-')
                    weeks.append({'year': int(year), 'week': int(week)})
                except ValueError:
                    pass
        return sorted(weeks, key=lambda x: (x['year'], x['week']), reverse=True)
    
    # ==================== 原始文本操作 ====================
    
    def _get_raw_text(self, year: int, week: int) -> tuple:
        """获取原始文本"""
        import json
        wb = self._get_workbook()
        if 'raw_text' not in wb.sheetnames:
            return '', {}
        
        ws = wb['raw_text']
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == year and ws.cell(row=row, column=2).value == week:
                raw_text = ws.cell(row=row, column=3).value or ''
                messages_str = ws.cell(row=row, column=4).value or '{}'
                try:
                    messages = json.loads(messages_str)
                except:
                    messages = {}
                return raw_text, messages
        return '', {}
    
    def _save_raw_text(self, year: int, week: int, raw_text: str, recommender_messages: Dict):
        """保存原始文本"""
        import json
        wb = self._get_workbook()
        
        if 'raw_text' not in wb.sheetnames:
            ws = wb.create_sheet('raw_text')
            ws.cell(row=1, column=1, value='year')
            ws.cell(row=1, column=2, value='week')
            ws.cell(row=1, column=3, value='raw_text')
            ws.cell(row=1, column=4, value='recommender_messages')
        else:
            ws = wb['raw_text']
        
        # 查找或添加
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == year and ws.cell(row=row, column=2).value == week:
                target_row = row
                break
        
        if not target_row:
            target_row = ws.max_row + 1
        
        ws.cell(row=target_row, column=1, value=year)
        ws.cell(row=target_row, column=2, value=week)
        ws.cell(row=target_row, column=3, value=raw_text)
        ws.cell(row=target_row, column=4, value=json.dumps(recommender_messages, ensure_ascii=False))
        
        wb.save(self.excel_path)
    
    # ==================== 推荐人统计操作 ====================
    
    def get_recommender_stats(self) -> List[Dict]:
        """获取所有推荐人统计"""
        import json
        wb = self._get_workbook()
        
        if 'recommender_stats' not in wb.sheetnames:
            return []
        
        ws = wb['recommender_stats']
        stats = []
        
        for row in range(2, ws.max_row + 1):
            stat = {}
            for i, col in enumerate(self.STATS_COLUMNS, 1):
                value = ws.cell(row=row, column=i).value
                if col == 'weekly_returns_json' and value:
                    try:
                        stat['weekly_returns'] = json.loads(value)
                    except:
                        stat['weekly_returns'] = []
                elif col == 'details_json' and value:
                    try:
                        stat['details'] = json.loads(value)
                    except:
                        stat['details'] = []
                elif col not in ['weekly_returns_json', 'details_json']:
                    stat[col] = value
            if stat.get('name'):
                stats.append(stat)
        
        return sorted(stats, key=lambda x: x.get('score', 0) or 0, reverse=True)
    
    def save_recommender_stats(self, stats_list: List[Dict]):
        """保存推荐人统计（全量覆盖）"""
        import json
        wb = self._get_workbook()
        
        # 删除旧sheet
        if 'recommender_stats' in wb.sheetnames:
            del wb['recommender_stats']
        
        # 创建新sheet
        ws = wb.create_sheet('recommender_stats', 0)
        
        # 写入表头
        for i, col in enumerate(self.STATS_COLUMNS, 1):
            ws.cell(row=1, column=i, value=col)
        
        # 写入数据
        for row_idx, stat in enumerate(stats_list, 2):
            for col_idx, col in enumerate(self.STATS_COLUMNS, 1):
                if col == 'weekly_returns_json':
                    value = json.dumps(stat.get('weekly_returns', []), ensure_ascii=False)
                elif col == 'details_json':
                    value = json.dumps(stat.get('details', []), ensure_ascii=False)
                else:
                    value = stat.get(col, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(self.excel_path)
    
    def delete_week(self, year: int, week: int) -> bool:
        """删除指定周数据"""
        wb = self._get_workbook()
        sheet_name = self._get_week_sheet_name(year, week)
        
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            wb.save(self.excel_path)
            return True
        return False


# 单例
_db_instance = None

def get_db():
    """根据配置返回数据源实例"""
    global _db_instance
    if _db_instance is None:
        if config.DATA_SOURCE == 'mongodb':
            from services.mongodb_db import MongoDB
            _db_instance = MongoDB()
            print(f"[DB] 使用 MongoDB 数据源: {config.MONGODB_DB}")
        else:
            _db_instance = ExcelDB()
            print(f"[DB] 使用 Excel 数据源: {config.EXCEL_FILE}")
    return _db_instance

